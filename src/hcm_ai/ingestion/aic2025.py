"""A lightweight adapter for AIC2025-style query sets and keyframe trees.

The adapter intentionally avoids importing video, OCR, or spreadsheet packages
until a caller asks for the corresponding operation.  This keeps ordinary unit
tests and CPU-only notebook setup fast.
"""

from __future__ import annotations

import json
import re
from collections.abc import Iterable, Mapping
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from ..contracts import FrameRecord, QueryRecord, TaskType


class OptionalDependencyError(RuntimeError):
    """Raised only when a caller explicitly requires an optional reader."""


_IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp", ".bmp"}
_VIDEO_EXTENSIONS = {".mp4", ".mkv", ".avi", ".mov", ".webm", ".m4v"}
_EXCEL_EXTENSIONS = {".xlsx", ".xlsm"}
_TASK_PATTERNS: tuple[tuple[TaskType, re.Pattern[str]], ...] = (
    (TaskType.TRAKE, re.compile(r"(?<![A-Z0-9])TRAKE(?![A-Z0-9])", re.IGNORECASE)),
    (TaskType.QA, re.compile(r"(?<![A-Z0-9])(?:VQA|QA)(?![A-Z0-9])", re.IGNORECASE)),
    (TaskType.KIS, re.compile(r"(?<![A-Z0-9])KIS(?![A-Z0-9])", re.IGNORECASE)),
)


def discover_aic2025_source(root: str | Path) -> tuple[str, Path]:
    """Find a conventional supplied source under one mounted dataset root.

    Discovery is intentionally conservative and deterministic. Authoritative
    frame metadata wins, followed by non-empty keyframes and raw videos.
    Non-standard layouts remain available through explicit CLI flags.
    """

    dataset_root = Path(root)
    if not dataset_root.is_dir():
        raise NotADirectoryError(dataset_root)

    metadata_candidates = (
        dataset_root / "frames.jsonl",
        dataset_root / "frames.json",
        dataset_root / "metadata" / "frames.jsonl",
        dataset_root / "metadata" / "frames.json",
        dataset_root / "metadata" / "frame_records.jsonl",
    )
    for candidate in metadata_candidates:
        if candidate.is_file() and candidate.stat().st_size > 0:
            return "frame_metadata", candidate

    lowered_name = dataset_root.name.casefold()
    keyframe_candidates = (
        *((dataset_root,) if lowered_name in {"keyframe", "keyframes"} else ()),
        *(dataset_root / name for name in ("keyframes", "Keyframes", "keyframe", "Keyframe")),
    )
    for candidate in keyframe_candidates:
        if candidate.is_dir() and any(
            path.is_file() and path.suffix.lower() in _IMAGE_EXTENSIONS
            for path in candidate.rglob("*")
        ):
            return "keyframes", candidate

    video_candidates = (
        *((dataset_root,) if lowered_name in {"video", "videos"} else ()),
        *(dataset_root / name for name in ("videos", "Videos", "video", "Video")),
    )
    for candidate in video_candidates:
        if candidate.is_dir() and any(
            path.is_file() and path.suffix.lower() in _VIDEO_EXTENSIONS
            for path in candidate.rglob("*")
        ):
            return "videos", candidate

    expected = ", ".join(
        str(path) for path in (*keyframe_candidates, *metadata_candidates, *video_candidates)
    )
    raise FileNotFoundError(
        f"could not discover keyframes, frame metadata, or videos under {dataset_root}; "
        f"checked: {expected}"
    )


def infer_task_type(value: str | Path | Iterable[str | Path], default: TaskType | None = None) -> TaskType:
    """Infer a task from a filename, directory names, or a collection of hints."""

    if isinstance(value, (str, Path)):
        values: Iterable[str | Path] = (value,)
    else:
        values = value
    combined = " ".join(str(item) for item in values)
    for task, pattern in _TASK_PATTERNS:
        if pattern.search(combined):
            return task
    if default is not None:
        return default
    raise ValueError(f"could not infer task type from {combined!r}")


def _read_query_text(path: Path) -> str:
    last_error: UnicodeError | None = None
    for encoding in ("utf-8-sig", "utf-8", "utf-16"):
        try:
            text = path.read_text(encoding=encoding)
            break
        except UnicodeError as exc:
            last_error = exc
    else:
        raise ValueError(f"could not decode query file {path}") from last_error

    lines = [line.strip() for line in text.splitlines() if line.strip()]
    if not lines:
        raise ValueError(f"query file {path} is empty")
    # A few query-set exports label their first content line.  Preserve all
    # natural-language content while removing the label itself.
    first = re.match(r"^(?:query|description|text)\s*[:=-]\s*(.+)$", lines[0], re.IGNORECASE)
    if first:
        lines[0] = first.group(1).strip()
    return "\n".join(lines)


def _task_hints(path: Path) -> list[str]:
    return [path.stem, *(parent.name for parent in path.parents)]


def parse_txt_query(path: str | Path, default_task: TaskType | None = None) -> QueryRecord:
    """Parse one ``.txt`` query and infer its AIC task from its filename/path."""

    source = Path(path)
    if source.suffix.lower() != ".txt":
        raise ValueError(f"expected a .txt query file, got {source}")
    task = infer_task_type(_task_hints(source), default=default_task)
    return QueryRecord(
        query_id=source.stem,
        text=_read_query_text(source),
        task=task,
        source_path=str(source),
        metadata={"source_format": "txt"},
    )


def _normalise_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").casefold())


def _cell_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _header_index(headers: list[Any], aliases: set[str]) -> int | None:
    for index, header in enumerate(headers):
        if _normalise_header(header) in aliases:
            return index
    return None


def parse_excel_queries(
    path: str | Path,
    *,
    default_task: TaskType | None = None,
    strict_optional_dependency: bool = False,
) -> list[QueryRecord]:
    """Read ``Query Name``, ``Description`` and optional ``Trans`` columns.

    When ``openpyxl`` is absent, ordinary discovery returns no rows so a text
    query set remains usable.  Set ``strict_optional_dependency=True`` to turn
    that condition into an actionable exception.
    """

    source = Path(path)
    if source.suffix.lower() not in _EXCEL_EXTENSIONS:
        raise ValueError(f"expected an .xlsx or .xlsm query workbook, got {source}")
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        if strict_optional_dependency:
            raise OptionalDependencyError(
                "openpyxl is required to read Excel query sets; install hcm-ai[dev]"
            ) from exc
        return []

    workbook = load_workbook(source, read_only=True, data_only=True)
    records: list[QueryRecord] = []
    try:
        for sheet in workbook.worksheets:
            rows = sheet.iter_rows(values_only=True)
            headers = next(rows, None)
            if headers is None:
                continue
            query_name_index = _header_index(list(headers), {"queryname", "queryid", "name"})
            description_index = _header_index(list(headers), {"description", "query", "text", "desc"})
            trans_index = _header_index(list(headers), {"trans", "translation", "translated", "english"})
            task_index = _header_index(list(headers), {"task", "tasktype", "type"})
            if query_name_index is None or description_index is None:
                continue

            for row_number, row in enumerate(rows, start=2):
                query_id = _cell_text(row[query_name_index] if query_name_index < len(row) else None)
                description = _cell_text(row[description_index] if description_index < len(row) else None)
                translation = _cell_text(row[trans_index] if trans_index is not None and trans_index < len(row) else None)
                if query_id is None or (description is None and translation is None):
                    continue
                task_hint = _cell_text(row[task_index] if task_index is not None and task_index < len(row) else None)
                task = infer_task_type(
                    [query_id, task_hint or "", sheet.title, source.name],
                    default=default_task or TaskType.KIS,
                )
                records.append(
                    QueryRecord(
                        query_id=query_id,
                        text=description or translation or "",
                        translated_text=translation,
                        task=task,
                        source_path=str(source),
                        metadata={"source_format": "excel", "sheet": sheet.title, "row": row_number},
                    )
                )
    finally:
        workbook.close()
    return records


def _merge_query_records(records: Iterable[QueryRecord]) -> list[QueryRecord]:
    """Deduplicate equivalent query exports while rejecting conflicting IDs."""

    merged: dict[str, QueryRecord] = {}
    for record in records:
        existing = merged.get(record.query_id)
        if existing is None:
            merged[record.query_id] = record
            continue
        if existing.task != record.task or existing.text != record.text:
            raise ValueError(f"conflicting query definitions for {record.query_id!r}")
        if existing.translated_text is None and record.translated_text is not None:
            merged[record.query_id] = existing.model_copy(update={"translated_text": record.translated_text})
    return [merged[key] for key in sorted(merged)]


def load_aic2025_queries(
    source: str | Path,
    *,
    default_task: TaskType | None = None,
    strict_optional_dependency: bool = False,
) -> list[QueryRecord]:
    """Load recursively discovered AIC query TXT/Excel files in stable order."""

    root = Path(source)
    if not root.exists():
        raise FileNotFoundError(root)
    if root.is_file():
        suffix = root.suffix.lower()
        if suffix == ".txt":
            return [parse_txt_query(root, default_task)]
        if suffix in _EXCEL_EXTENSIONS:
            return parse_excel_queries(
                root,
                default_task=default_task,
                strict_optional_dependency=strict_optional_dependency,
            )
        raise ValueError(f"unsupported query source {root}")

    records: list[QueryRecord] = []
    for path in sorted(root.rglob("*.txt")):
        records.append(parse_txt_query(path, default_task))
    for extension in sorted(_EXCEL_EXTENSIONS):
        for path in sorted(root.rglob(f"*{extension}")):
            records.extend(
                parse_excel_queries(
                    path,
                    default_task=default_task,
                    strict_optional_dependency=strict_optional_dependency,
                )
            )
    return _merge_query_records(records)


def _frame_number_from_name(path: Path) -> int | None:
    match = re.search(r"(\d+)(?!.*\d)", path.stem)
    return int(match.group(1)) if match else None


def _video_id_from_keyframe_path(path: Path, root: Path) -> str:
    relative = path.relative_to(root)
    if len(relative.parts) > 1:
        return relative.parts[-2]
    # Flat trees are uncommon but remain usable when a filename begins with a
    # video prefix such as L01_V001_000123.jpg.
    match = re.match(r"(.+?)(?:[_-]F?\d+)?$", path.stem)
    return match.group(1) if match else path.stem


def build_keyframe_manifest(
    keyframes_root: str | Path,
    *,
    fps_by_video: Mapping[str, float] | None = None,
    default_fps: float = 25.0,
) -> list[FrameRecord]:
    """Create a deterministic ``FrameRecord`` manifest from a keyframe tree.

    AIC exports commonly encode a source frame number in the image filename.
    This helper converts it to seconds using per-video FPS when provided, with
    an explicit default-FPS provenance marker otherwise.  Existing official
    metadata can be converted to ``FrameRecord`` directly instead.
    """

    root = Path(keyframes_root)
    if default_fps <= 0.0:
        raise ValueError("default_fps must be positive")
    if not root.is_dir():
        raise NotADirectoryError(root)
    fps_lookup = dict(fps_by_video or {})
    records: list[FrameRecord] = []
    for image_path in sorted(path for path in root.rglob("*") if path.suffix.lower() in _IMAGE_EXTENSIONS):
        video_id = _video_id_from_keyframe_path(image_path, root)
        frame_number = _frame_number_from_name(image_path)
        fps = float(fps_lookup.get(video_id, default_fps))
        if fps <= 0.0:
            raise ValueError(f"fps for video {video_id!r} must be positive")
        timestamp = (frame_number or 0) / fps
        frame_id = image_path.stem if image_path.stem.startswith(video_id) else f"{video_id}_{image_path.stem}"
        records.append(
            FrameRecord(
                frame_id=frame_id,
                video_id=video_id,
                timestamp=timestamp,
                image_path=str(image_path),
                frame_number=frame_number,
                metadata={
                    "manifest_source": "keyframe_tree",
                    "fps": fps,
                    "timestamp_source": "frame_number/fps" if frame_number is not None else "default_zero",
                },
            )
        )
    return records


def load_frame_records(path: str | Path) -> list[FrameRecord]:
    """Read JSON or JSONL frame records from legacy media-info exports."""

    source = Path(path)
    if source.suffix.lower() == ".jsonl":
        rows = [json.loads(line) for line in source.read_text(encoding="utf-8").splitlines() if line.strip()]
    elif source.suffix.lower() == ".json":
        payload = json.loads(source.read_text(encoding="utf-8"))
        if isinstance(payload, Mapping):
            rows = payload.get("frames", payload.get("records", []))
        else:
            rows = payload
    else:
        raise ValueError(f"expected .json or .jsonl frame metadata, got {source}")
    if not isinstance(rows, list):
        raise ValueError(f"frame metadata {source} must contain a list of records")
    return [FrameRecord.model_validate(row) for row in rows]


@dataclass(frozen=True)
class AIC2025Adapter:
    """Convenience adapter rooted at a Drive-mounted AIC2025 data directory."""

    root: Path

    def __init__(self, root: str | Path) -> None:
        object.__setattr__(self, "root", Path(root))

    def load_queries(
        self,
        relative_path: str | Path = ".",
        *,
        default_task: TaskType | None = None,
        strict_optional_dependency: bool = False,
    ) -> list[QueryRecord]:
        return load_aic2025_queries(
            self.root / relative_path,
            default_task=default_task,
            strict_optional_dependency=strict_optional_dependency,
        )

    def build_keyframe_manifest(
        self,
        relative_path: str | Path = "keyframes",
        *,
        fps_by_video: Mapping[str, float] | None = None,
        default_fps: float = 25.0,
    ) -> list[FrameRecord]:
        return build_keyframe_manifest(
            self.root / relative_path,
            fps_by_video=fps_by_video,
            default_fps=default_fps,
        )


__all__ = [
    "AIC2025Adapter",
    "OptionalDependencyError",
    "build_keyframe_manifest",
    "discover_aic2025_source",
    "infer_task_type",
    "load_aic2025_queries",
    "load_frame_records",
    "parse_excel_queries",
    "parse_txt_query",
]
